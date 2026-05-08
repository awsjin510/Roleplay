"""Generate Roleplay 產品開發時間彙整表 (.xlsx)

依據 git 歷史，將實際開發過的功能填入 Excel 表格。
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path


HEADERS = [
    "編號",
    "功能名稱",
    "模組分類",
    "功能描述",
    "開始日期",
    "完成日期",
    "實際工時（人日）",
    "產出物 / 涉及檔案",
    "備註 / 心得",
]

# 依據 git log 整理的實際開發紀錄
ROWS = [
    ("F-001", "轉盤生成器基礎架構", "前端",
     "Vanilla JS 多欄位轉盤動畫引擎，隨機抽取練習情境",
     "2026/03/13", "2026/03/13", 1.0,
     "index.html, js/spinner.js, js/app.js",
     "首次提交建立 MVP 框架"),
    ("F-002", "視覺設計（Glassmorphism）", "設計",
     "玻璃擬態 + 漸層配色 + 轉盤特效",
     "2026/03/13", "2026/03/13", 0.5,
     "css/main.css, css/spinner.css",
     "與基礎架構同日完成，UI 品質直接拉到 demo-ready"),
    ("F-003", "業務人物 + 產品線轉盤欄", "前端+資料",
     "新增業務人物（9 人）與產品線（9 種）兩個轉盤欄位",
     "2026/03/13", "2026/03/13", 0.5,
     "js/data.js, js/spinner.js",
     "讓情境更貼近真實業務場景"),
    ("F-004", "情境資料庫擴充", "資料",
     "15 產業 × 12 客戶角色 × 40+ 客戶情境 × 20 難題",
     "2026/03/13", "2026/03/31", 1.5,
     "js/data.js",
     "資料量決定情境豐富度，期間多次微調（移除 Joanna 等）"),
    ("F-005", "AI 角色扮演器（React 版）", "前端+AI",
     "React + JSX artifact，含對話 UI、評分、教練覆盤",
     "2026/04/03", "2026/04/03", 2.0,
     "sales-roleplay-trainer.jsx",
     "進階版主體，整合 Babel 即時編譯"),
    ("F-006", "AI 對話器（HTML 版）+ API Key 管理", "AI 整合",
     "輕量 HTML 版本，使用者自帶 Claude API key",
     "2026/04/03", "2026/04/03", 2.0,
     "roleplay-ai.html",
     "讓任何人都能即開即用，不需後端"),
    ("F-007", "Anthropic Browser Header 修正", "AI 整合",
     "瀏覽器直連 Claude API 需要的特殊 header",
     "2026/04/03", "2026/04/03", 0.5,
     "roleplay-ai.html",
     "卡了一陣子才發現是 anthropic-dangerous-direct-browser-access"),
    ("F-008", "移除 Recharts 依賴 + 按鈕樣式", "前端",
     "拔掉外部 CDN，避免載入失敗；統一按鈕樣式",
     "2026/04/03", "2026/04/03", 0.5,
     "sales-roleplay-trainer.jsx",
     "減少外部依賴提高穩定性"),
    ("F-009", "教練面板加寬 + 回合計數器", "前端",
     "覆盤面板顯示空間不夠，調寬並修可見度",
     "2026/04/03", "2026/04/03", 0.5,
     "sales-roleplay-trainer.jsx",
     "review failure 修復"),
    ("F-010", "上下導覽按鈕 + 欄寬調整", "前端",
     "每個轉盤欄加上下手動切換按鈕；難度欄加寬",
     "2026/04/04", "2026/04/04", 0.5,
     "js/spinner.js, css/spinner.css",
     "使用者回饋希望可以手動微調抽到的選項"),
    ("F-011", "AI Trainer 與 Spinner 資料同步", "資料",
     "兩個版本共用同一份 data.js 避免資料漂移",
     "2026/04/03", "2026/04/03", 0.5,
     "js/data.js",
     "重構降低維護成本"),
    ("F-012", "雲端情境 + 公有雲環境轉盤", "前端+資料",
     "新增 AWS / GCP / Azure 環境維度，補齊雲端銷售情境",
     "2026/04/07", "2026/04/07", 1.5,
     "index.html, roleplay-ai.html, js/data.js",
     "本專案核心競爭力：雲端業務情境練習"),
    ("F-013", "條件式雲端環境顯示", "前端",
     "只在抽到雲端服務產品時才顯示雲端環境轉盤",
     "2026/04/07", "2026/04/07", 0.5,
     "index.html, js/spinner.js",
     "避免非雲端產品出現不相關欄位"),
    ("F-014", "轉盤動畫優化 + 落定閃光", "設計",
     "縮短動畫時間提升節奏感，加入落定瞬間閃光特效",
     "2026/04/08", "2026/04/08", 0.5,
     "css/spinner.css, js/spinner.js",
     "perf 優化 + 視覺回饋"),
]


def style_header(ws):
    header_fill = PatternFill("solid", fgColor="305496")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    ws.row_dimensions[1].height = 30


def style_body(ws, n_rows):
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="top", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1, max_col=len(HEADERS)):
        for cell in row:
            cell.border = border
            cell.alignment = wrap


def add_total_row(ws, n_rows):
    total_row = n_rows + 2
    ws.cell(row=total_row, column=6, value="總計").font = Font(bold=True)
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=7,
                         value=f"=SUM(G2:G{n_rows + 1})")
    total_cell.font = Font(bold=True, color="C00000")
    total_cell.fill = PatternFill("solid", fgColor="FFF2CC")


def set_column_widths(ws):
    widths = [8, 22, 12, 32, 12, 12, 14, 30, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_validation(ws, n_rows):
    dv = DataValidation(
        type="list",
        formula1='"前端,AI 整合,前端+AI,資料,前端+資料,設計,部署,後端"',
        allow_blank=True,
    )
    dv.add(f"C2:C{n_rows + 50}")
    ws.add_data_validation(dv)


def build_summary_sheet(wb):
    ws = wb.create_sheet("模組工時分析")
    ws["A1"] = "模組分類"
    ws["B1"] = "累計工時（人日）"
    ws["C1"] = "佔比"
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center")

    modules = [
        "前端", "AI 整合", "前端+AI", "資料", "前端+資料", "設計", "部署", "後端"
    ]
    for i, m in enumerate(modules, start=2):
        ws.cell(row=i, column=1, value=m)
        ws.cell(row=i, column=2,
                value=f'=SUMIF(\'開發時間彙整\'!C:C,A{i},\'開發時間彙整\'!G:G)')
        ws.cell(row=i, column=3,
                value=f"=IFERROR(B{i}/SUM($B$2:$B${len(modules)+1}),0)")
        ws.cell(row=i, column=3).number_format = "0.0%"

    total_row = len(modules) + 2
    ws.cell(row=total_row, column=1, value="總計").font = Font(bold=True)
    ws.cell(row=total_row, column=2,
            value=f"=SUM(B2:B{total_row - 1})").font = Font(bold=True)
    ws.cell(row=total_row, column=2).fill = PatternFill("solid", fgColor="FFF2CC")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "開發時間彙整"

    ws.append(HEADERS)
    for row in ROWS:
        ws.append(list(row))

    style_header(ws)
    style_body(ws, len(ROWS))
    set_column_widths(ws)
    add_total_row(ws, len(ROWS))
    add_validation(ws, len(ROWS))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(ROWS) + 1}"

    build_summary_sheet(wb)

    out_path = Path(__file__).resolve().parent.parent / "Roleplay-開發時間彙整.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
