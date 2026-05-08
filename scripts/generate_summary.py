"""
Generate project feature summary Excel.

Output: docs/專案功能彙整表.xlsx
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FEATURES = [
    # (功能名稱, 模組分類, 開發工時(人天), 備註)
    ("業務角色扮演轉盤首頁",       "前端 / 抽籤模組 (index.html)", 1.0, "六欄轉盤：業務 / 產品 / 產業 / 角色 / 情境 / 難度，提供入口頁。"),
    ("轉盤動畫元件 SpinnerWheel", "前端 / 抽籤模組 (js/spinner.js)", 1.5, "支援單欄轉動、上下步進、點擊選取，使用 CSS transform 與 transitionend 完成動畫。"),
    ("全部轉動 / 重置功能",       "前端 / 抽籤模組 (js/app.js)",   0.5, "依序觸發 6 顆轉盤；產品為「雲端服務」時動態顯示雲端環境第七顆轉盤。"),
    ("情境資料字典",              "前端 / 資料層 (js/data.js)",     0.5, "9 位業務、9 種產品、15 種產業、12 種角色、40 種情境、20 種難度、5 種雲端環境。"),
    ("響應式 UI / 主題樣式",      "前端 / 樣式 (css/main.css, spinner.css)", 1.0, "漸層標題、Pulse Glow 按鈕、Noto Sans TC 中文字型，支援桌機與手機。"),

    ("AI 對話練習主頁面",         "前端 / AI 練習 (roleplay-ai.html)", 2.0, "React (CDN) 單檔應用；包含設定、對話、檢討三大階段切換。"),
    ("情境設定 SetupPhase",       "前端 / AI 練習 (roleplay-ai.html)", 1.0, "老虎機式 SlotColumn 抽籤；產品為雲端 / 資安時動態加開對應欄位。"),
    ("AI 對話 ChatPhase",         "前端 / AI 練習 (roleplay-ai.html)", 2.5, "整合 Anthropic Messages API、即時對話、回合計數、提示與計時器。"),
    ("信任 / 興趣 / 購買意願指標", "前端 / 指標儀表板",                 1.0, "CircularGauge 三圈儀表 + SimpleLineChart 折線圖呈現對話過程指標變化。"),
    ("檢討回饋 ReviewPhase",      "前端 / AI 練習 (roleplay-ai.html)", 1.0, "對話結束後呼叫 Claude 產生評分、優點、改進建議與下次練習方向。"),
    ("語音輸入 (Web Speech API)", "前端 / 互動體驗",                   0.5, "支援 zh-TW SpeechRecognition，按鈕切換錄音狀態。"),
    ("提示卡片 / Toast 訊息",     "前端 / 互動體驗",                   0.3, "隨機抽取 HINTS 內建話術提示，輔助新手業務。"),
    ("角色 / 產業 / 難度 Profile", "前端 / Prompt 工程",               1.5, "roleProfiles、industryProfiles、difficultyBehaviors 等字典組合 systemPrompt。"),

    ("RAG 知識庫管理後台",        "前端 / 管理 (admin.html)",          2.0, "Google 帳號登入、文件清單、上傳 / 刪除文件介面。"),
    ("文件解析 (PDF/DOCX/XLSX/TXT)", "前端 / 文件處理",                1.5, "整合 pdf.js、mammoth、SheetJS，將原始檔轉為純文字後分塊。"),
    ("Google Drive 匯入",          "前端 / 文件處理",                  1.0, "Google Picker 選檔，支援 Google Docs / Sheets 自動匯出格式。"),
    ("文字切塊 chunkText",         "前端 / 文件處理",                  0.5, "依段落合併並保留 100 字元 overlap，最大 800 字元 chunk。"),
    ("產品分類自動推測",           "前端 / 文件處理",                  0.3, "依檔名關鍵字推測對應產品，方便上傳分類。"),

    ("Cloudflare Worker 反向代理", "後端 / Edge (worker.js)",           1.0, "/chat 代理 Anthropic Messages API；/embed 呼叫 Workers AI 產生向量。"),
    ("X-App-Token 驗證 / CORS",   "後端 / 安全",                       0.3, "requireAppToken middleware 與 OPTIONS preflight 處理。"),
    ("Workers AI 向量嵌入",       "後端 / 向量化",                     0.5, "@cf/baai/bge-base-en-v1.5（768 維），供 RAG 相似度查詢使用。"),
    ("Wrangler 部署設定",         "DevOps (wrangler.toml)",            0.3, "compatibility_date、AI binding 與 Secret 設定（ANTHROPIC_API_KEY、APP_TOKEN）。"),

    ("Firebase Firestore 整合",   "後端 / 資料庫",                     1.0, "練習紀錄 practice_records 與 RAG rag_docs / rag_chunks 兩大集合。"),
    ("匿名登入 / 練習紀錄同步",   "後端 / 資料庫",                     0.5, "signInAnonymously，雲端與 localStorage 雙寫，雲端不可用時降級。"),
    ("RAG 相似度檢索 findNearest","後端 / 向量檢索",                   1.0, "依 scenario 組合查詢字串 -> 取得 embedding -> Firestore 取最近 chunks。"),

    ("API Key 設定 Modal",        "前端 / 設定",                       0.3, "Anthropic API Key 儲存於 localStorage，支援即時驗證。"),
    ("使用者名稱 / 歷史紀錄",     "前端 / 設定",                       0.3, "localStorage 儲存最多 200 筆紀錄並可刪除。"),
    ("舊版 React JSX 原型",       "歷史 / 參考檔 (sales-roleplay-trainer.jsx)", 1.0, "首版單檔 React 原型，現由 roleplay-ai.html 取代，保留作參考。"),
]


def auto_width(ws, max_width=80):
    for col_idx, column in enumerate(ws.columns, start=1):
        letter = get_column_letter(col_idx)
        longest = 0
        for cell in column:
            if cell.value is None:
                continue
            for line in str(cell.value).split("\n"):
                # Chinese characters take ~2 width, others ~1
                width = sum(2 if ord(c) > 127 else 1 for c in line)
                longest = max(longest, width)
        ws.column_dimensions[letter].width = min(max_width, longest + 4)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "專案功能彙整"

    title_font = Font(name="Microsoft JhengHei", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="6366F1")
    body_font = Font(name="Microsoft JhengHei", size=10)
    total_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="111827")
    total_fill = PatternFill("solid", fgColor="E0E7FF")
    thin = Side(style="thin", color="C7D2FE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="left")
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")

    ws.merge_cells("A1:E1")
    title = ws["A1"]
    title.value = "業務角色扮演訓練系統 - 功能彙整表"
    title.font = title_font
    title.fill = title_fill
    title.alignment = center
    ws.row_dimensions[1].height = 32

    ws["A2"] = "產出日期：2026-05-08"
    ws["A2"].font = Font(name="Microsoft JhengHei", size=9, italic=True, color="6B7280")
    ws.merge_cells("A2:E2")

    headers = ["#", "功能名稱", "模組分類", "開發工時 (人天)", "備註"]
    header_row = 4
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 26

    for idx, (name, module, hours, note) in enumerate(FEATURES, start=1):
        row = header_row + idx
        values = [idx, name, module, hours, note]
        aligns = [center, wrap_center, wrap_center, center, wrap_center]
        for col, (v, align) in enumerate(zip(values, aligns), start=1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = body_font
            cell.alignment = align
            cell.border = border
            if col == 4:
                cell.number_format = "0.0"
        ws.row_dimensions[row].height = 38
        if idx % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="F5F3FF")

    total_row = header_row + len(FEATURES) + 1
    ws.cell(row=total_row, column=1, value="").border = border
    ws.cell(row=total_row, column=2, value="總計").alignment = center
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
    total_hours = sum(f[2] for f in FEATURES)
    ws.cell(row=total_row, column=4, value=total_hours).number_format = "0.0"
    ws.cell(row=total_row, column=5, value=f"共 {len(FEATURES)} 項功能")
    for col in range(1, 6):
        c = ws.cell(row=total_row, column=col)
        c.font = total_font
        c.fill = total_fill
        c.border = border
        if c.alignment.horizontal is None:
            c.alignment = center
    ws.row_dimensions[total_row].height = 30

    ws.freeze_panes = "A5"

    auto_width(ws)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 70

    out_dir = Path(__file__).resolve().parent.parent / "docs"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "專案功能彙整表.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
